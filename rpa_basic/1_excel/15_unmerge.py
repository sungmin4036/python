from openpyxl import load_workbook
wb = load_workbook("sample_unmerge.xlsx")
ws = wb.active


ws.unmerge_cells("B2:D2")