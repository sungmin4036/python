from openpyxl import load_workbook
from openpyxl.styles import Font
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

#A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1 행의 높이를 50 으로 설정
ws.row_dimensions[1].height= 50

#스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 색 빨갛게, 이탤릭, 두껍게
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 폰트를 Arial로 설정, 가운데에 줄 긑기
c1.font = Font(color="0000FF", size=20, underline="single") # 글짜크기 20, 밑줄


wb.save("sample_style.xlsx")