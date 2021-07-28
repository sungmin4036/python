from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() # 샐운 Sheet 기본 이름으로 생성
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" #RGB 형태로 값을 넣어주면 탭 색상 변경

# Sheet, mySheet, YourSheet
ws1 = wb.create_sheet("YourtSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번쨰 indeex 에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복스
new_ws["A1"] = "Test" # A1 셀에 Test 내용 넣기
target = wb.copy_worksheet(new_ws) # new_ws == NewSheet 의 페이지를 복사
target.title = "Copied Sheet" #

wb.save("sample.xlsx")