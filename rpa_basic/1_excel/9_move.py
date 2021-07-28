from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

ws.move_range("B1:C11", rows=0, cols=1) # B1부터 C11 까지 row는 그대로, cols만 한칸 이동

ws["B1"].value = "국어"

wb.save("sample_korea.xlsx")