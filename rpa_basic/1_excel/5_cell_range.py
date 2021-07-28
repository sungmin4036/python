from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string
wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] #영어 column 만 가지고 오기
# # print(col_B)

# for cell in col_B:
#     print(cell.value)
# wb.save("sample.xlsx")

# col_range = ws["B:C"] # B부터 C까지 (영어, 수학 column) 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1] # 1번쨰 row 만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2번째 줄에서 6번쨰 줄까지 가지고 오기
# for rows in row_range:
#     for cell in rows:
#          print(cell.value, end= " ")
#     print()

# row_range = ws[2:ws.max_row] # 2번쨰 줄부터 마지막줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end= " ")
#         # print(cell.coordinate, end=" ") # 셀의 좌표
#         xy = coordinate_from_string(cell.coordinate) # A/10,  AZ/250 이런식으로 끊어줌
#         # print(xy, end=" ")
#         print(xy[0], end=" ") # A
#         print(xy[1], end=" ") # 1
#     print()


#전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)


#전체 columns
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[1].value)

# for row in ws.iter_rows(): # 전체 row
#     print(row)

# for column in ws.iter_cols(): # 전체 column
#     print(column[0].value, end=" ")

# 1번쨰 줄부터 5번째 줄, 2번쨰 열부터 3번쨰 열까지
# for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
#     print(row[0].value, row[1].value) # 수학, 영어

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)
wb.save("sample.xlsx")